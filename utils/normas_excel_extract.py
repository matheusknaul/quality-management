from openpyxl import load_workbook
import openpyxl

def gerar_excel(lista):

    wb = openpyxl.Workbook()
    ws = wb.active

    #codigo, descricao, ano_escopo, status, ultima_verificacao, observacao

    ws['A1'] = 'Código'
    ws['B1'] = 'Descrição'
    ws['C1'] = 'Ano da norma no escopo'
    ws['D1'] = 'Status'
    ws['E1'] = 'Última verificação'
    ws['F1'] = 'Observação'

    for item in range(len(lista)):
        ws[f'A{item + 2}'] = lista[item][0]
        ws[f'B{item + 2}'] = lista[item][1]
        ws[f'C{item + 2}'] = lista[item][2]
        ws[f'D{item + 2}'] = lista[item][3]
        ws[f'E{item + 2}'] = lista[item][4]
        ws[f'F{item + 2}'] = ""
    
    file_name = 'utils/normas_verificacao.xlsx'

    wb.save(file_name)

    print('Extract de normas feito com sucesso!')