from types import NoneType
from database.models.norma import Norma

"""
ATENÇÃO: Utilizar esse código somente quando quiser resetar o banco de dados.

Esse código serve, caso você precise adicionar registros no banco de dados no shell.

"""

from openpyxl import load_workbook
import openpyxl
import re

_dataframepath_ = 'dataframe.xlsx'

def get_dado_excel(planilha_excel):
    wb = load_workbook(planilha_excel)
    sheet = wb['Sheet1']

    total_linhas = sheet.max_row

    next_linha = []
    for row in range(2, total_linhas + 1):
        print(f'número da linha {row}')
        codigo_norma = str(sheet[f'A{row}'].value) + " " + str(sheet[f'B{row}'].value)
        tag_norma = sheet[f'A{row}'].value
        number_norma = sheet[f'B{row}'].value
        ano_da_norma = sheet[f'E{row}'].value
        if not tag_norma:
            tag_norma = "NULL"
        if not number_norma:
            number_norma = "NULL"
        if not ano_da_norma:
            ano_da_norma = "NULL"
        status = sheet[f'F{row}'].value
        partes_da_norma = sheet[f'D{row}'].value # --------> talvez mudar
        if partes_da_norma:
            partes = re.findall(r'\d', sheet[f'D{row}'].value)
            print(f"esse daqui é a lista de partes gerada do findall{partes}")
            while partes:
                primeira_parte = partes.pop(0)
                new_codigo_norma = codigo_norma + "-" + primeira_parte
                new_norma = Norma.create(codigo=new_codigo_norma, descricao='', ano_norma=ano_da_norma, situacao=status, tag_main=tag_norma, part_main=primeira_parte, number_main=number_norma)
                print(f"esse daqui é a lista de partes após o pop{partes}")
        else:
            new_norma = Norma.create(codigo=codigo_norma, descricao='', ano_norma=ano_da_norma, situacao=status, tag_main=tag_norma, part_main="NULL", number_main=number_norma)

get_dado_excel(_dataframepath_)